from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import *
from datetime import date, datetime
from django.db.models import Q
from django.utils import timezone
from notificationapp.models import Notification
from adminapp.models import (
    tbl_patient,
    tbl_medical_record,
    tbl_prescription,
    tbl_doctor
)
from patientapp.models import tbl_appointment


def doctorindex(request):
    login_id = request.session.get('login_id')
    unread_count = 0
    doctor_name = "Doctor"
    total_patients = 0
    todays_appointments = 0
    high_risk_cases = 0
    pending_reports = 0
    total_deliveries = 0
    recent_patients = []
    
    if login_id:
        try:
            unread_count = Notification.objects.filter(user_type='Doctor', user_id=str(login_id), is_read=False).count()
        except Exception:
            pass
        
        try:
            # Get the doctor details
            doctor = tbl_doctor.objects.get(login_id=login_id)
            doctor_name = doctor.doctor_name
            
            # Get total patients assigned to this doctor
            total_patients = tbl_patient.objects.filter(doctor_id=doctor).count()
            
            # Get today's active appointments (accepted today or rescheduled to today)
            today = timezone.localdate()
            todays_appointments = tbl_appointment.objects.filter(
                doctor_id=doctor
            ).filter(
                Q(status='Accepted', appointment_date=today) |
                Q(status='Rescheduled', reschedule_date=today)
            ).count()
            
            # Get high risk cases (patients with complications or specific trimester)
            # This is a placeholder - adjust based on your actual criteria
            high_risk_cases = tbl_patient.objects.filter(doctor_id=doctor, current_trimester=3).count()
            
            # Get pending reports count (medical records without doctor notes)
            from adminapp.models import tbl_medical_record
            pending_reports = tbl_medical_record.objects.filter(
                patient_id__doctor_id=doctor,
                doctor_note__isnull=True
            ).count()
            
            # Get total delivered cases (patients with delivery_date not null)
            total_deliveries = tbl_patient.objects.filter(doctor_id=doctor, delivery_date__isnull=False).count()
            
            # Get recent patients (last 5)
            patients = tbl_patient.objects.filter(doctor_id=doctor).order_by('-patient_id')[:5]
            
            for patient in patients:
                # Get last visit
                last_visit_record = tbl_visit_history.objects.filter(patient_id=patient).order_by('-visit_date').first()
                last_visit = last_visit_record.visit_date if last_visit_record else None
                
                recent_patients.append({
                    'patient_id': patient.patient_id,
                    'patient_name': patient.patient_name,
                    'current_trimester': patient.current_trimester or 'N/A',
                    'last_visit': last_visit.strftime('%d %b %Y') if last_visit else 'No visits yet',
                    'status': 'Active',
                    'status_class': 'active'
                })
                
        except Exception as e:
            pass
    
    context = {
        'unread_count': unread_count,
        'doctor_name': doctor_name,
        'total_patients': total_patients,
        'todays_appointments': todays_appointments,
        'high_risk_cases': high_risk_cases,
        'pending_reports': pending_reports,
        'total_deliveries': total_deliveries,
        'recent_patients': recent_patients
    }
    
    return render(request, 'Doctor/index.html', context)

def doctor_logout(request):
    request.session.flush()
    messages.success(request, 'You have been logged out successfully.')
    return redirect('guestapp:guestindex')

def patientindex(request):
    return render(request, 'Patient/index.html')
def doctor_patients(request):
    
    login_id = request.session.get('login_id')

    if not login_id:
        patients = tbl_patient.objects.none()
    else:
        try:
            doctor = tbl_doctor.objects.get(login_id=login_id)
            patients = tbl_patient.objects.filter(doctor_id=doctor)
        except Exception:
            patients = tbl_patient.objects.none()

    return render(request, 'Doctor/viewpatients.html', {
        'patients': patients
    })


def patient_info(request, patient_id):
    """Display patient basic information"""
    try:
        patient = get_object_or_404(tbl_patient, patient_id=patient_id)
        
        # Calculate current month of pregnancy
        current_month = None
        if patient.last_cycle_date:
            from datetime import date as date_class
            days_diff = (date_class.today() - patient.last_cycle_date).days
            current_month = int(days_diff / 30.44)  # Convert to months (30.44 = average days per month)
        
        return render(request, 'Doctor/patient_info.html', {
            'patient': patient,
            'current_month': current_month
        })
    except Exception as e:
        messages.error(request, 'Error loading patient information.')
        return redirect('doctorapp:doctor_patients')


def delivery_due_soon(request):
    """Display patients with delivery due within 30 days"""
    from datetime import date, timedelta
    
    login_id = request.session.get('login_id')
    patients_list = []
    
    if not login_id:
        return redirect('guestapp:guestindex')
    
    try:
        doctor = tbl_doctor.objects.get(login_id=login_id)
        today = date.today()
        thirty_days_from_now = today + timedelta(days=30)
        
        # Get patients assigned to this doctor with EDD within 30 days
        patients = tbl_patient.objects.filter(
            doctor_id=doctor,
            edd_date__isnull=False,
            edd_date__gte=today,
            edd_date__lte=thirty_days_from_now
        ).order_by('edd_date')
        
        for patient in patients:
            days_until_due = (patient.edd_date - today).days
            
            # Determine urgency
            if days_until_due <= 7:
                urgency_class = 'urgent'
            elif days_until_due <= 14:
                urgency_class = 'soon'
            else:
                urgency_class = 'upcoming'
            
            patients_list.append({
                'patient_id': patient.patient_id,
                'patient_name': patient.patient_name,
                'edd_date': patient.edd_date.strftime('%d %b %Y'),
                'current_trimester': patient.current_trimester,
                'age': patient.age,
                'phone': patient.phone,
                'days_until_due': days_until_due,
                'urgency_class': urgency_class
            })
    
    except Exception as e:
        messages.error(request, 'Error loading delivery due soon list.')
    
    return render(request, 'Doctor/delivery_due_soon.html', {
        'patients': patients_list
    })


def generate_report(request):
    """Display patient list for report generation"""
    login_id = request.session.get('login_id')
    
    if not login_id:
        messages.error(request, 'Please login to access this page.')
        return redirect('guestapp:guestindex')
    
    try:
        doctor = tbl_doctor.objects.get(login_id=login_id)
        patients = tbl_patient.objects.filter(doctor_id=doctor).order_by('patient_name')
        
        # Count patients by trimester (current_trimester is IntegerField: 1, 2, 3)
        total_patients = patients.count()
        first_trimester = patients.filter(current_trimester=1).count()
        second_trimester = patients.filter(current_trimester=2).count()
        third_trimester = patients.filter(current_trimester=3).count()
        
        context = {
            'patients': patients,
            'total_patients': total_patients,
            'first_trimester': first_trimester,
            'second_trimester': second_trimester,
            'third_trimester': third_trimester
        }
        
        return render(request, 'Doctor/generate_report.html', context)
        
    except tbl_doctor.DoesNotExist:
        messages.error(request, 'Doctor record not found. Please contact administrator.')
        return redirect('doctorapp:doctorindex')
    except Exception as e:
        messages.error(request, f'Error loading patient data: {str(e)}')
        return redirect('doctorapp:doctorindex')


def export_patients_excel(request):
    """Generate and download Excel report of patients"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    login_id = request.session.get('login_id')
    
    if not login_id:
        return redirect('guestapp:guestindex')
    
    try:
        doctor = tbl_doctor.objects.get(login_id=login_id)
        patients = tbl_patient.objects.filter(doctor_id=doctor).order_by('patient_name')
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patient Report"
        
        # Define header style
        header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Patient Report - Dr. {doctor.doctor_name}"
        title_cell.font = Font(bold=True, size=14, color="E91E63")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add generation date
        ws.merge_cells('A2:G2')
        date_cell = ws['A2']
        date_cell.value = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        headers = ['Patient ID', 'Patient Name', 'Age', 'Phone', 'Email', 
                   'Trimester', 'EDD']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add patient data
        row_num = 5
        for patient in patients:
            # Convert trimester number to name
            trimester_map = {1: 'First', 2: 'Second', 3: 'Third'}
            trimester_display = trimester_map.get(patient.current_trimester, 'N/A') if patient.current_trimester else 'N/A'
            
            ws.cell(row=row_num, column=1, value=patient.patient_id)
            ws.cell(row=row_num, column=2, value=patient.patient_name)
            ws.cell(row=row_num, column=3, value=patient.age)
            ws.cell(row=row_num, column=4, value=patient.phone or 'N/A')
            ws.cell(row=row_num, column=5, value=patient.email or 'N/A')
            ws.cell(row=row_num, column=6, value=trimester_display)
            ws.cell(row=row_num, column=7, value=str(patient.edd_date) if patient.edd_date else 'N/A')
            row_num += 1
        
        # Adjust column widths
        column_widths = [12, 25, 8, 15, 30, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Add summary at the bottom
        summary_row = row_num + 2
        ws.cell(row=summary_row, column=1, value="Summary:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value="Total Patients:")
        ws.cell(row=summary_row + 1, column=2, value=patients.count())
        
        ws.cell(row=summary_row + 2, column=1, value="First Trimester:")
        ws.cell(row=summary_row + 2, column=2, value=patients.filter(current_trimester=1).count())
        
        ws.cell(row=summary_row + 3, column=1, value="Second Trimester:")
        ws.cell(row=summary_row + 3, column=2, value=patients.filter(current_trimester=2).count())
        
        ws.cell(row=summary_row + 4, column=1, value="Third Trimester:")
        ws.cell(row=summary_row + 4, column=2, value=patients.filter(current_trimester=3).count())
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Patient_Report_{doctor.doctor_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating Excel report: {str(e)}')
        return redirect('doctorapp:generate_report')


def patient_profile(request, patient_id):
    from datetime import date
    from adminapp.utils import (
        ensure_patient_lock_status,
        can_add_prescription,
        can_update_trimester,
        is_patient_profile_readonly,
        check_role_can_add_notes,
        log_status_change
    )
   
    patient = get_object_or_404(tbl_patient, patient_id=patient_id)
    
    # Check if patient profile is locked or unlock has expired
    try:
        is_locked_status = ensure_patient_lock_status(patient)
    except Exception:
        is_locked_status = patient.profile_lock_status == 'locked'
    
    # Additional check for unlock expiry
    if patient.unlock_end_date and date.today() > patient.unlock_end_date:
        is_locked_status = True

    # Check if profile is read-only based on status
    profile_readonly = is_patient_profile_readonly(patient)
    if patient.status == 'Miscarriage':
        profile_readonly = False

 
    if request.method == 'POST':
        # Get current user role
        login_id = request.session.get('login_id')
        try:
            from guestapp.models import tbl_login
            login_obj = tbl_login.objects.get(login_id=login_id)
            user_role = login_obj.role
        except Exception:
            user_role = 'Patient'
        
        action = request.POST.get('action')
        
        # Check role authorization to update status
        if action == 'update_status':
            if user_role not in ['Doctor', 'Admin']:
                messages.error(request, 'You do not have permission to update patient status.')
                return redirect('doctorapp:patient_profile', patient_id=patient_id)
            
            old_status = patient.status
            new_status = request.POST.get('new_status')
            status_notes = request.POST.get('status_notes', '')

            if not new_status:
                messages.info(request, 'No status change selected. Patient remains Active.')
                return redirect('doctorapp:patient_profile', patient_id=patient_id)

            valid_status_values = [choice[0] for choice in tbl_patient.STATUS_CHOICES]
            if new_status not in valid_status_values:
                messages.error(request, 'Invalid status selected.')
                return redirect('doctorapp:patient_profile', patient_id=patient_id)

            if old_status == new_status:
                messages.info(request, f'Patient is already in {new_status} status.')
                return redirect('doctorapp:patient_profile', patient_id=patient_id)
            
            # Only Doctor can add miscarriage or transfer notes
            if new_status == 'Miscarriage':
                if status_notes and user_role == 'Doctor':
                    patient.miscarriage_notes = status_notes
                # Capture miscarriage date when status is set
                if not patient.miscarriage_date:
                    patient.miscarriage_date = date.today()
            
            if new_status == 'Emergency' and status_notes and user_role == 'Doctor':
                patient.emergency_notes = status_notes
            
            if new_status == 'Transferred' and status_notes and user_role == 'Doctor':
                patient.transfer_reason = status_notes
                transfer_date_raw = request.POST.get('transfer_date')
                if transfer_date_raw:
                    try:
                        patient.transfer_date = datetime.strptime(transfer_date_raw, '%Y-%m-%d').date()
                    except Exception:
                        pass
                transfer_summary = request.POST.get('transfer_summary', '')
                if transfer_summary:
                    patient.transfer_summary = transfer_summary
                # Immediately lock profile when transferred
                patient.profile_lock_status = 'locked'
                patient.lock_start_date = date.today()
            
            patient.status = new_status
            patient.save()
            
            log_status_change(patient, old_status, new_status, status_notes)
            messages.success(request, f'Patient status updated to {new_status}')
        
        # Prevent prescription and notes if patient profile is locked/unlock expired
        elif is_locked_status or profile_readonly:
            messages.warning(request, 'Cannot modify records for this patient due to status or lock.')
            return redirect('doctorapp:patient_profile', patient_id=patient_id)
        
        # Prevent prescription if status restricts it
        elif action == 'add_prescription':
            if patient.status != 'Miscarriage' and not can_add_prescription(patient):
                messages.error(request, f'Cannot add prescription for patient with status: {patient.status}')
                return redirect('doctorapp:patient_profile', patient_id=patient_id)
            
            login_id = request.session.get('login_id')
            if login_id:
                doctor = get_object_or_404(tbl_doctor, login_id=login_id)
                
                diagnosis = request.POST.get('diagnosis')
                medicines = request.POST.get('medicines')
                dosage = request.POST.get('dosage')
                additional_notes = request.POST.get('additional_notes')
                
                tbl_prescription.objects.create(
                    patient=patient,
                    doctor=doctor,
                    diagnosis=diagnosis,
                    medicines=medicines,
                    dosage=dosage,
                    additional_notes=additional_notes
                )
                
                # Notify patient and admin
                try:
                    Notification.objects.create(
                        user_type='Patient',
                        user_id=str(patient.login_id.login_id),
                        message=f'Dr. {doctor.doctor_name} added a prescription for {diagnosis}'
                    )
                    # Notify all admins
                    from guestapp.models import tbl_login
                    admin_users = tbl_login.objects.filter(role='Admin')
                    for admin in admin_users:
                        Notification.objects.create(
                            user_type='Admin',
                            user_id=str(admin.login_id),
                            message=f'Dr. {doctor.doctor_name} added prescription for patient {patient.patient_name}'
                        )
                    messages.success(request, 'Prescription added successfully!')
                    print("✓ Notifications created successfully")
                except Exception as e:
                    print(f"✗ Notification error: {e}")
        
        # Add doctor notes for medical records
        elif action != 'add_prescription':
            record_id = request.POST.get('record_id')
            doctor_note = request.POST.get('doctor_note')

            if record_id and doctor_note:
                record = get_object_or_404(
                    tbl_medical_record,
                    record_id=record_id,
                    patient_id=patient   
                )
                record.doctor_note = doctor_note
                record.save()

        return redirect('doctorapp:patient_profile', patient_id=patient_id)

    # 🔹 Calculate current pregnancy month
    if patient.last_cycle_date:
        delta_days = (date.today() - patient.last_cycle_date).days
        patient.current_month = (delta_days // 30) + 1
        # Cap at 9 months maximum for pregnancy
        if patient.current_month > 9:
            patient.current_month = 9
    else:
        patient.current_month = None

    # 🔹 Medical records
    medical_records = tbl_medical_record.objects.filter(
        patient_id=patient   # ✅ FIXED
    ).order_by('-record_date')

    # 🔹 Prescriptions
    prescriptions = tbl_prescription.objects.filter(
        patient_id=patient   # ✅ FIXED
    ).order_by('-prescription_date')

    # 🔹 Latest visit history
    latest_visit = tbl_visit_history.objects.filter(
        patient_id=patient   # ✅ FIXED
    ).order_by('-visit_date').first()

    # 🔹 Delivery details (if already entered)
    delivery = tbl_delivery_details.objects.filter(patient_id=patient).first()

    # Get user role for template checks
    login_id = request.session.get('login_id')
    user_role = 'Patient'
    try:
        from guestapp.models import tbl_login
        login_obj = tbl_login.objects.get(login_id=login_id)
        user_role = login_obj.role
    except Exception:
        pass

    context = {
        'patient': patient,
        'medical_records': medical_records,
        'prescriptions': prescriptions,
        'latest_visit': latest_visit,
        'delivery': delivery,
        'is_locked': is_locked_status,
        'profile_readonly': profile_readonly,
        'can_add_prescription': can_add_prescription(patient),
        'can_update_trimester': can_update_trimester(patient),
        'user_role': user_role,
        'can_update_status': user_role in ['Doctor', 'Admin'],
        'patient_status_choices': tbl_patient.STATUS_CHOICES,
    }

    return render(
        request,
        'Doctor/patientprofile.html',
        context
    )




from django.shortcuts import render
from datetime import date, datetime
from django.db.models import Q
from patientapp.models import tbl_appointment


def doctor_appointments_by_date(request):
    doctor_id = request.session.get('doctor_id')

    
    if not doctor_id:
        login_id = request.session.get('login_id')
        if login_id:
            try:
                from adminapp.models import tbl_doctor
                doctor = tbl_doctor.objects.filter(login_id=login_id).first()
                if doctor:
                    doctor_id = doctor.doctor_id
            except Exception:
                doctor_id = None

    
    selected_date_str = request.GET.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except Exception:
            selected_date = date.today()
    else:
        selected_date = date.today()

    
    appointments = tbl_appointment.objects.filter(
        doctor_id=doctor_id,
        status__in=['Accepted', 'Rescheduled']
    ).filter(
        Q(appointment_date=selected_date) | Q(reschedule_date=selected_date)
    )
    
    # Mark which appointments already have visits
    for appointment in appointments:
        appointment.has_visit = tbl_visit_history.objects.filter(
            appointment_id=appointment.appointment_id
        ).exists()

    return render(request, 'Doctor/acceptedpatients.html', {
        'appointments': appointments,
        'selected_date': selected_date
    })

from django.shortcuts import render, redirect
from datetime import date
from adminapp.models import tbl_appointment, tbl_doctor,tbl_patient
from .models import tbl_visit_history

from django.shortcuts import get_object_or_404

def add_visit(request, appointment_id):
    appointment = get_object_or_404(tbl_appointment, appointment_id=appointment_id)
    patient = appointment.patient_id
    doctor = appointment.doctor_id

    if request.method == 'POST':
        tbl_visit_history.objects.create(
            patient_id=patient,
            doctor_id=doctor,
            appointment_id=appointment,
            visit_date=date.today(),
            details=request.POST.get('details'),
            weight=request.POST.get('weight'),
            blood_pressure=request.POST.get('blood_pressure'),
            health_status=request.POST.get('health_status'),
            next_visit_date=request.POST.get('next_visit_date')
        )
        messages.success(request, f'Visit details added successfully for {patient.patient_name}!')
        return redirect('doctorapp:doctor_appointments_by_date')

    return render(request, 'Doctor/add_visit.html', {
        'patient': patient,
        'appointment': appointment
    })

from django.shortcuts import render, get_object_or_404
from .models import tbl_visit_history
from adminapp.models import tbl_appointment


def doctor_visits(request):
    login_id = request.session.get('login_id')
    
    if not login_id:
        visits = tbl_visit_history.objects.none()
    else:
        try:
            doctor = tbl_doctor.objects.get(login_id=login_id)
            visits = tbl_visit_history.objects.filter(doctor_id=doctor).order_by('-visit_date')
        except Exception:
            visits = tbl_visit_history.objects.none()
    
    return render(request, 'Doctor/doctor_visits.html', {
        'visits': visits
    })


def view_visit(request, appointment_id):
    visit = get_object_or_404(
        tbl_visit_history,
        appointment_id=appointment_id
    )

    return render(request, 'Doctor/view_visit.html', {
        'visit': visit
    })

from django.shortcuts import render, redirect, get_object_or_404
from datetime import date

from adminapp.models import tbl_patient, tbl_appointment,tbl_prescription,tbl_doctor



def add_prescription(request, patient_id, appointment_id):
    # Logged-in doctor
    doctor = get_object_or_404(
        tbl_doctor,
        login_id=request.session.get('login_id')
    )

    patient = get_object_or_404(tbl_patient, patient_id=patient_id)
    appointment = get_object_or_404(tbl_appointment, appointment_id=appointment_id)

    if request.method == 'POST':
        diagnosis = request.POST.get('diagnosis')
        medicines = request.POST.get('medicines')
        dosage = request.POST.get('dosage')
        additional_notes = request.POST.get('additional_notes')

        tbl_prescription.objects.create(
            patient=patient,
            doctor=doctor,
            appointment=appointment,
            diagnosis=diagnosis,
            medicines=medicines,
            dosage=dosage,
            additional_notes=additional_notes
        )
        
        # Send prescription email to patient
        from adminapp.utils import send_email
        subject = "Your Prescription Has Been Uploaded"
        message = f"""Dear {patient.patient_name},

Dr. {doctor.doctor_name} has uploaded a prescription for you.

Prescription Details:
Diagnosis: {diagnosis}
Medicines: {medicines}
Dosage: {dosage}
Additional Notes: {additional_notes if additional_notes else 'None'}

Please visit your health portal to view and download the complete prescription.

Best regards,
Bloom Moms Care Team"""
        send_email(subject, message, patient.email)
        messages.success(request, f'Prescription added successfully for {patient.patient_name}!')

        return redirect(
            'doctorapp:patient_profile',
            patient_id=patient.patient_id
        )

    return render(request, 'Doctor/add_prescription.html', {
        'patient': patient,
        'appointment': appointment
    })


from django.shortcuts import render, redirect, get_object_or_404
from adminapp.models import tbl_patient, tbl_doctor
from .models import tbl_delivery_details

def add_delivery_details(request, patient_id):
    """
    Doctor adds or updates delivery details for a patient.
    """
    # Check if doctor is logged in via session
    login_id = request.session.get('login_id')
    if not login_id:
        return redirect('guestapp:login')

    # Get user role
    try:
        from guestapp.models import tbl_login
        login_obj = tbl_login.objects.get(login_id=login_id)
        user_role = login_obj.role
        
        # Only Doctor and Admin can add delivery details
        if user_role not in ['Doctor', 'Admin']:
            messages.error(request, 'You do not have permission to add delivery details.')
            return redirect('guestapp:guestindex')
    except Exception:
        messages.error(request, 'Authentication error. Please login again.')
        return redirect('guestapp:login')

    # Get the patient object
    patient = get_object_or_404(tbl_patient, patient_id=patient_id)

    # Block adding delivery details for transferred or miscarriage patients
    if patient.status == 'Transferred':
        messages.error(request, 'Cannot add delivery details for transferred patients.')
        return redirect('doctorapp:patient_profile', patient_id=patient_id)
    
    if patient.status == 'Miscarriage':
        messages.error(request, 'Cannot add delivery details for miscarriage cases.')
        return redirect('doctorapp:patient_profile', patient_id=patient_id)

    # Get logged-in doctor from session
    doctor = get_object_or_404(tbl_doctor, login_id=login_id)

    # Check if delivery details already exist
    delivery = tbl_delivery_details.objects.filter(patient=patient).first()

    if request.method == 'POST':
        delivery_date_raw = request.POST.get('delivery_date')
        delivery_type = request.POST.get('delivery_type')
        baby_weight = request.POST.get('baby_weight')
        baby_condition = request.POST.get('baby_condition')
        mother_condition = request.POST.get('mother_condition')
        remarks = request.POST.get('remarks')

        delivery_date = None
        try:
            delivery_date = datetime.strptime(delivery_date_raw, '%Y-%m-%d').date()
        except Exception:
            delivery_date = None

        if delivery:  # update existing record
            delivery.delivery_date = delivery_date
            delivery.delivery_type = delivery_type
            delivery.baby_weight = baby_weight
            delivery.baby_condition = baby_condition
            delivery.mother_condition = mother_condition
            delivery.remarks = remarks
            delivery.save()
        else:  # create new record
            tbl_delivery_details.objects.create(
                patient=patient,
                doctor=doctor,
                delivery_date=delivery_date,
                delivery_type=delivery_type,
                baby_weight=baby_weight,
                baby_condition=baby_condition,
                mother_condition=mother_condition,
                remarks=remarks
            )
        
        # Notify patient about delivery
        Notification.objects.create(
            user_type='Patient',
            user_id=str(patient.login_id.login_id),
            message=f'Your delivery details have been recorded by Dr. {doctor.doctor_name}'
        )

        # Store delivery date on patient and reset lock until grace period expires.
        if delivery_date:
            patient.delivery_date = delivery_date
            patient.profile_lock_status = 'unlocked'
            patient.lock_start_date = None
            patient.status = 'Delivered'  # Update status to Delivered
            patient.save(update_fields=['delivery_date', 'profile_lock_status', 'lock_start_date', 'status'])

            messages.success(request, f'Delivery details recorded successfully for {patient.patient_name}!')

            # Immediately enforce lock if delivery was long ago.
            try:
                from adminapp.utils import ensure_patient_lock_status
                is_now_locked = ensure_patient_lock_status(patient)
            except Exception:
                pass

        # Redirect back to patient profile
        return redirect('doctorapp:patient_profile', patient_id=patient_id)

    context = {
        'patient': patient,
        'delivery': delivery,
        'user_role': user_role,
    }
    return render(request, 'Doctor/add_delivery_details.html', context)


def view_delivery_details(request, patient_id):
    """
    Display existing delivery details for a patient. Redirects to the add/update
    form if no delivery record exists yet.
    """
    login_id = request.session.get('login_id')
    if not login_id:
        return redirect('guestapp:login')

    patient = get_object_or_404(tbl_patient, patient_id=patient_id)
    delivery = tbl_delivery_details.objects.filter(patient=patient).first()

    if not delivery:
        return redirect('doctorapp:add_delivery_details', patient_id=patient_id)

    context = {
        'patient': patient,
        'delivery': delivery
    }
    return render(request, 'Doctor/view_delivery_details.html', context)


    